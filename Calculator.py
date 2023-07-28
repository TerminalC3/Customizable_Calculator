from tkinter import *
from tkinter import colorchooser
import openpyxl as xl
import math


def button_press(num):

    global equation_text
    global equation_variable

    equation_text = display.get()

    if num == 'π':
        num = 22/7

        if equation_text != "":
            equation_text = float(equation_text) * float(num)
            equation_variable.set(equation_text)

        else:
            equation_text = float(num)
            equation_variable.set(equation_text)

    else:
        equation_text = str(equation_text) + str(num)
        equation_variable.set(equation_text)

    display.icursor(len(str(equation_text)))


def equals1(event):

    global equation_text

    equation_text = display.get()

    error_variable.set("")

    try:
        total = str(eval(equation_text))

        equation_variable.set(total)

        equation_text = total
        display.icursor(len(str(equation_text)))

    except SyntaxError:
        error_variable.set("Syntax Error")

    except ZeroDivisionError:
        error_variable.set("Zero Division Error")


def equals():

    global equation_text

    equation_text = display.get()

    error_variable.set("")

    try:
        total = str(eval(equation_text))

        equation_variable.set(total)

        equation_text = total
        display.icursor(len(str(equation_text)))

    except SyntaxError:
        error_variable.set("Syntax Error")

    except ZeroDivisionError:
        error_variable.set("Zero Division Error")


def sq():
    global equation_text

    square = float(equation_text) * float(equation_text)

    equation_variable.set(square)
    equation_text = square
    display.icursor(len(str(equation_text)))


def sqr_root():
    global equation_text

    squr_rt = math.sqrt(float(equation_text))

    equation_variable.set(squr_rt)
    equation_text = squr_rt
    display.icursor(len(str(equation_text)))


def delete():
    global equation_text

    equation_text = display.get()

    equation_text = equation_text[0 : -1]

    equation_variable.set(equation_text)


def delete1():
    global equation_text

    equation_text = display.get()

    equation_text = equation_text[0 : -1]

    equation_variable.set(equation_text)


def copy():
    r = Tk()
    r.withdraw()
    r.clipboard_clear()
    r.clipboard_append(display.get())
    r.update()
    r.destroy()


def text_box_color():
    global color

    color = colorchooser.askcolor()

    display.config(bg = color[1])

    sheet.cell(1,1).value = color[1]
    wb.save('colors.xlsx')


def text_box_fg_color():
    global color

    color = colorchooser.askcolor()

    display.config(fg =color[1])

    sheet.cell(1, 2).value = color[1]
    wb.save('colors.xlsx')


def background_color():
    global color

    color = colorchooser.askcolor()

    window.config(bg = color[1])
    fill_label.config(bg = color[1])
    fill_label2.config(bg = color[1])
    error_label.config(bg = color[1])

    sheet.cell(1, 3).value = color[1]
    wb.save('colors.xlsx')


def button_bg():
    global color

    color = colorchooser.askcolor()

    button1.config(bg=color[1])
    button2.config(bg=color[1])
    button3.config(bg=color[1])
    button4.config(bg=color[1])
    button5.config(bg=color[1])
    button6.config(bg=color[1])
    button7.config(bg=color[1])
    button8.config(bg=color[1])
    button9.config(bg=color[1])
    button0.config(bg=color[1])
    equals.config(bg=color[1])
    decimal.config(bg=color[1])
    delete.config(bg=color[1])
    clear.config(bg=color[1])
    options.config(bg=color[1])
    equals.config(bg=color[1])
    plus.config(bg=color[1])
    minus.config(bg=color[1])
    divide.config(bg=color[1])
    multiply.config(bg=color[1])
    close.config(bg=color[1])
    copy.config(bg=color[1])
    square.config(bg=color[1])
    square_root.config(bg=color[1])
    pi.config(bg=color[1])
    delete1.config(bg=color[1])

    sheet.cell(1, 4).value = color[1]
    wb.save('colors.xlsx')


def button_fg():
    global color

    color = colorchooser.askcolor()

    button1.config(fg=color[1])
    button2.config(fg = color[1])
    button3.config(fg=color[1])
    button4.config(fg=color[1])
    button5.config(fg=color[1])
    button6.config(fg=color[1])
    button7.config(fg=color[1])
    button8.config(fg=color[1])
    button9.config(fg=color[1])
    button0.config(fg=color[1])
    equals.config(fg=color[1])
    decimal.config(fg=color[1])
    delete.config(fg=color[1])
    clear.config(fg=color[1])
    options.config(fg=color[1])
    equals.config(fg=color[1])
    plus.config(fg=color[1])
    minus.config(fg=color[1])
    divide.config(fg=color[1])
    multiply.config(fg=color[1])
    close.config(fg=color[1])
    copy.config(fg=color[1])
    square.config(fg=color[1])
    square_root.config(fg=color[1])
    pi.config(fg=color[1])
    delete1.config(fg=color[1])

    sheet.cell(1, 5).value = color[1]
    wb.save('colors.xlsx')


def error_text_color():
    global color

    color = colorchooser.askcolor()

    error_label.config(fg = color[1])

    sheet.cell(1, 6).value = color[1]
    wb.save('colors.xlsx')


def hover_button_color():
    global color
    global hover_color

    color = colorchooser.askcolor()
    sheet.cell(1, 7).value = color[1]
    wb.save('colors.xlsx')

    hover_color = sheet.cell(1, 7).value


def on_enter(event):
    button1.config(bg = hover_color)


def on_leave(event):
    button1.config(bg = sheet.cell(1, 4).value)


def on_enter1(event):
    clear.config(bg = hover_color)


def on_leave1(event):
    clear.config(bg = sheet.cell(1, 4).value)


def on_enter2(event):
    close.config(bg = hover_color)


def on_leave2(event):
    close.config(bg = sheet.cell(1, 4).value)


def on_enter3(event):
    multiply.config(bg = hover_color)


def on_leave3(event):
    multiply.config(bg = sheet.cell(1, 4).value)


def on_enter4(event):
    divide.config(bg = hover_color)


def on_leave4(event):
    divide.config(bg = sheet.cell(1, 4).value)


def on_enter5(event):
    minus.config(bg = hover_color)


def on_leave5(event):
    minus.config(bg = sheet.cell(1, 4).value)


def on_enter6(event):
    plus.config(bg = hover_color)


def on_leave6(event):
    plus.config(bg = sheet.cell(1, 4).value)


def on_enter7(event):
    equals.config(bg = hover_color)


def on_leave7(event):
    equals.config(bg = sheet.cell(1, 4).value)


def on_enter8(event):
    button2.config(bg = hover_color)


def on_leave8(event):
    button2.config(bg = sheet.cell(1, 4).value)


def on_enter9(event):
    button3.config(bg = hover_color)


def on_leave9(event):
    button3.config(bg = sheet.cell(1, 4).value)


def on_enter0(event):
    button4.config(bg = hover_color)


def on_leave0(event):
    button4.config(bg = sheet.cell(1, 4).value)


def on_enter10(event):
    button5.config(bg = hover_color)


def on_leave10(event):
    button5.config(bg = sheet.cell(1, 4).value)


def on_enter11(event):
    button6.config(bg = hover_color)


def on_leave11(event):
    button6.config(bg = sheet.cell(1, 4).value)


def on_enter12(event):
    button7.config(bg = hover_color)


def on_leave12(event):
    button7.config(bg = sheet.cell(1, 4).value)


def on_enter13(event):
    button8.config(bg = hover_color)


def on_leave13(event):
    button8.config(bg = sheet.cell(1, 4).value)


def on_enter14(event):
    button9.config(bg = hover_color)


def on_leave14(event):
    button9.config(bg = sheet.cell(1, 4).value)


def on_enter15(event):
    button0.config(bg = hover_color)


def on_leave15(event):
    button0.config(bg = sheet.cell(1, 4).value)


def on_enter16(event):
    pi.config(bg = hover_color)


def on_leave16(event):
    pi.config(bg = sheet.cell(1, 4).value)


def on_enter17(event):
    square.config(bg = hover_color)


def on_leave17(event):
    square.config(bg = sheet.cell(1, 4).value)


def on_enter18(event):
    square_root.config(bg = hover_color)


def on_leave18(event):
    square_root.config(bg = sheet.cell(1, 4).value)


def on_enter19(event):
    decimal.config(bg = hover_color)


def on_leave19(event):
    decimal.config(bg = sheet.cell(1, 4).value)


def on_enter20(event):
    delete.config(bg = hover_color)


def on_leave20(event):
    delete.config(bg = sheet.cell(1, 4).value)


def on_enter21(event):
    copy.config(bg = hover_color)


def on_leave21(event):
    copy.config(bg = sheet.cell(1, 4).value)


def on_enter22(event):
    clear.config(bg = hover_color)


def on_leave22(event):
    clear.config(bg = sheet.cell(1, 4).value)


def on_enter23(event):
    delete1.config(bg = hover_color)


def on_leave23(event):
    delete1.config(bg = sheet.cell(1, 4).value)


def on_enter24(event):
    close.config(bg = hover_color)


def on_leave24(event):
    close.config(bg = sheet.cell(1, 4).value)


def on_enter25(event):
    options.config(bg = hover_color)


def on_leave25(event):
    options.config(bg = sheet.cell(1, 4).value)


def options():
    option_window = Tk()
    option_window.title("Options")
    option_window.geometry("450x450")
    option_window.config(bg="red")

    filler_label1 = Label(option_window, bg = "red")
    filler_label2 = Label(option_window, bg = "red")
    filler_label3 = Label(option_window, bg = "red")
    filler_label4 = Label(option_window, bg = "red")
    filler_label5 = Label(option_window, bg = "red")
    filler_label6 = Label(option_window, bg = "red")
    filler_label7 = Label(option_window, bg="red")

    filler_label6.pack()

    text_box_color_button = Button(option_window, text = "Change text box color", font = ("consolas, 15"), command = text_box_color, bg = "orange")
    text_box_color_button.pack()

    filler_label1.pack()

    text_box_fg_color_button = Button(option_window, text="Change text color", font=("Arial, 15"), command= text_box_fg_color, bg = "orange")
    text_box_fg_color_button.pack()

    filler_label2.pack()

    background_color_button = Button(option_window, text = "Change background color", font = ("Arial, 15"), command = background_color, bg = "orange")
    background_color_button.pack()

    filler_label3.pack()

    button_bg_color_button = Button(option_window, text = "Change button background color", font = ("Arial, 15"), command = button_bg, bg = "orange")
    button_bg_color_button.pack()

    filler_label4.pack()

    button_fg_color_button = Button(option_window, text = "Change button foreground color", font = ("Arial, 15"), command = button_fg, bg = "orange")
    button_fg_color_button.pack()

    filler_label5.pack()

    button_hover_color_button = Button(option_window, text="Change button hover color", font=("Arial, 15"), command= hover_button_color, bg="orange")
    button_hover_color_button.pack()

    filler_label7.pack()

    error_text_color_button = Button(option_window, text="Change error text color", font=("Arial, 15"), command= error_text_color, bg = "orange")
    error_text_color_button.pack()

    option_window.mainloop()


def clear():

    global equation_text

    equation_variable.set("")
    equation_text = ""


def close():
    window.destroy()


wb = xl.load_workbook('colors.xlsx')
sheet = wb["Sheet1"]

window = Tk()
window.geometry("500x530")
window.title("Calculator")
window.iconphoto(False, PhotoImage(file='Calc.png'))
window.config(bg = sheet.cell(1, 3).value)

hover_color = sheet.cell(1, 7).value

error_variable = StringVar()
equation_variable = StringVar()
equation_text = ""

fill_label = Label(window, bg = sheet.cell(1, 3).value)
fill_label.pack()

display = Entry(window, textvariable = equation_variable, font = ("consolas", 30, "bold"), relief = "sunken", bg = sheet.cell(1, 1).value, fg = sheet.cell(1, 2).value)
display.pack()
display.focus()

fill_label2 = Label(window, bg = sheet.cell(1, 3).value, height = 4)
fill_label2.pack()

frame = Frame(window)
frame.pack()

window.bind("<Return>", equals1)

button1 = Button(frame, text = "1", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(1), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button1.grid(row = 2, column = 0)

button2 = Button(frame, text = "2", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(2), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button2.grid(row = 2, column = 1)

button3 = Button(frame, text = "3", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(3), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button3.grid(row = 2, column = 2)

button4 = Button(frame, text = "4", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(4), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button4.grid(row = 1, column = 0)

button5 = Button(frame, text = "5", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(5), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button5.grid(row = 1, column = 1)

button6 = Button(frame, text = "6", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(6), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button6.grid(row = 1, column = 2)

button7 = Button(frame, text = "7", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(7), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button7.grid(row = 0, column = 0)

button8 = Button(frame, text = "8", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(8), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button8.grid(row = 0, column = 1)

button9 = Button(frame, text = "9", font = ("consolas", 30, "bold"), width = 4, command = lambda :button_press(9), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button9.grid(row = 0, column = 2)

button0 = Button(frame, text = "0", font = ("consolas", 30, "bold"), width = 4, command = lambda : button_press(0), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
button0.grid(row = 3, column = 1)

equals = Button(frame, text = "=", font = ("consolas", 30, "bold"), width = 4, command = equals, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
equals.grid(row = 3, column = 2)

plus = Button(frame, text = "+", font = ("consolas", 30, "bold"), width = 4, command = lambda : button_press('+'), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
plus.grid(row = 0 , column = 3)

minus = Button(frame, text = "-", font = ("consolas", 30, "bold"), width = 4, command = lambda : button_press('-'), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
minus.grid(row = 1 , column = 3)

divide = Button(frame, text = "÷", font = ("consolas", 30, "bold"), width = 4, command = lambda : button_press('/'), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
divide.grid(row = 2 , column = 3)

multiply = Button(frame, text = "×", font = ("consolas", 30, "bold"), width = 4, command = lambda : button_press('*'), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
multiply.grid(row = 3 , column = 3)

pi = Button(frame, text = "π", font = ("consolas", 30, "bold"), width = 4, command = lambda : button_press('π'), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
pi.grid(row = 1 , column = 4)

square = Button(frame, text = "x²", font = ("consolas", 30, "bold"), width = 4, command = sq, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
square.grid(row = 2 , column = 4)

square_root = Button(frame, text = "√x", font = ("consolas", 30, "bold"), width = 4, command = sqr_root, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
square_root.grid(row = 3 , column = 4)

decimal = Button(frame, text = ".", font = ("consolas", 15, "bold"), width = 8, command = lambda : button_press('.'), bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
decimal.grid(row = 4 , column = 3)

delete = Button(frame, text = "⌫", font = ("consolas", 30, "bold"), width = 4, height = 1, command = delete, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
delete.grid(row = 0, column = 4)

copy = Button(frame, text = "Copy", font = ("consolas", 30, "bold"), width = 4, height = 1, command = copy, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
copy.grid(row = 3, column = 0)

clear = Button(frame, text = "clear", font = ("consolas", 15, "bold"), width = 8, command = clear, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
clear.grid(row = 4, column = 2)

options = Button(frame, text = "options", font = ("consolas", 15, "bold"), width = 8, command = options, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
options.grid(row = 4, column = 0)

close = Button(frame, text = "close", font = ("consolas", 15, "bold"), width = 8, command = close, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
close.grid(row = 4, column = 4)

delete1 = Button(frame, text = "⌫", font = ("consolas", 15, "bold"), width = 8, height = 1, command = delete1, bg = sheet.cell(1, 4).value, fg = sheet.cell(1, 5).value)
delete1.grid(row = 4, column = 1)

error_label = Label(window, textvariable = error_variable, font = ("consolas", 15, "bold"), width = 20, bg = sheet.cell(1, 3).value, fg = sheet.cell(1, 6).value)
error_label.pack()

button1.bind("<Enter>", on_enter)
button1.bind("<Leave>", on_leave)

clear.bind("<Enter>", on_enter1)
clear.bind("<Leave>", on_leave1)

close.bind("<Enter>", on_enter2)
close.bind("<Leave>", on_leave2)

multiply.bind("<Enter>", on_enter3)
multiply.bind("<Leave>", on_leave3)

divide.bind("<Enter>", on_enter4)
divide.bind("<Leave>", on_leave4)

minus.bind("<Enter>", on_enter5)
minus.bind("<Leave>", on_leave5)

plus.bind("<Enter>", on_enter6)
plus.bind("<Leave>", on_leave6)

equals.bind("<Enter>", on_enter7)
equals.bind("<Leave>", on_leave7)

button2.bind("<Enter>", on_enter8)
button2.bind("<Leave>", on_leave8)

button3.bind("<Enter>", on_enter9)
button3.bind("<Leave>", on_leave9)

button4.bind("<Enter>", on_enter0)
button4.bind("<Leave>", on_leave0)

button5.bind("<Enter>", on_enter10)
button5.bind("<Leave>", on_leave10)

button6.bind("<Enter>", on_enter11)
button6.bind("<Leave>", on_leave11)

button7.bind("<Enter>", on_enter12)
button7.bind("<Leave>", on_leave12)

button8.bind("<Enter>", on_enter13)
button8.bind("<Leave>", on_leave13)

button9.bind("<Enter>", on_enter14)
button9.bind("<Leave>", on_leave14)

button0.bind("<Enter>", on_enter15)
button0.bind("<Leave>", on_leave15)

pi.bind("<Enter>", on_enter16)
pi.bind("<Leave>", on_leave16)

square.bind("<Enter>", on_enter17)
square.bind("<Leave>", on_leave17)

square_root.bind("<Enter>", on_enter18)
square_root.bind("<Leave>", on_leave18)

decimal.bind("<Enter>", on_enter19)
decimal.bind("<Leave>", on_leave19)

delete.bind("<Enter>", on_enter20)
delete.bind("<Leave>", on_leave20)

copy.bind("<Enter>", on_enter21)
copy.bind("<Leave>", on_leave21)

clear.bind("<Enter>", on_enter22)
clear.bind("<Leave>", on_leave22)

delete1.bind("<Enter>", on_enter23)
delete1.bind("<Leave>", on_leave23)

close.bind("<Enter>", on_enter24)
close.bind("<Leave>", on_leave24)

options.bind("<Enter>", on_enter25)
options.bind("<Leave>", on_leave25)

window.mainloop()
